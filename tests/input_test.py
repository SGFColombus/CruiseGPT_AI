from datetime import datetime
import requests
from typing import Dict, Any
import pandas as pd
import os
import numpy as np
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import json
import openpyxl
from openpyxl.styles import Alignment


# Create the data directory if it doesn't exist
data_dir = os.path.join(os.path.dirname(__file__), "data")
os.makedirs(data_dir, exist_ok=True)

# Define the Excel file path
EXCEL_FILE = os.path.join(data_dir, "input_test.xlsx")

# Define all attributes that need scoring
json_attributes = [
    "embarkationPort",
    "disembarkationPort",
    "destinations",
    "minDuration",
    "maxDuration",
    "minSailStartDate",
    "maxSailStartDate",
    "minSailEndDate",
    "maxSailEndDate",
    "maxPrice",
    "minPrice",
    "round_trip",
    "ignore_destinations",
    "price_discount",
    "is_expedition",
]

# Define the columns
columns = [
    "User Message",
    "Expected Features Extraction",
    "Actual Features Extraction",
]

# Add score columns for each JSON attribute
score_columns = [f"Score for {attr}" for attr in json_attributes]
columns.extend(score_columns)

# Add final score column
columns.append("Final Score")

# Define sample data
sample_data = [
    {
        "User Message": "Give me cruises that are going to Asia",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Asia"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Great choice! You\'ll be heading to Vancouver.",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find all cruises that last more than 10 days",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": 11,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises lasting more than 10 days",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find cruises that depart from Hong Kong",
        "Expected Features Extraction": (
            '[{"embarkationPort": ["Hong Kong"],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises departing from Hong Kong",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "I want cruises that arrive at Singapore",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": ["Singapore"],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises arriving at Singapore",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find cruises where the itinerary includes Gibraltar",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Gibraltar"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises that visit Gibraltar",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Are there cruises starting after January 1, 2025",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2025-01-01","maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises starting after January 1, 2025",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Provide me with cruises that end in Singapore and last exactly 11 days",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": ["Singapore"],'
            '"destinations": [],"minDuration": 11,"maxDuration": 11,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are 11-day cruises ending in Singapore",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find cruises departing from Hong Kong and traveling to Asia",
        "Expected Features Extraction": (
            '[{"embarkationPort": ["Hong Kong"],"disembarkationPort": [],'
            '"destinations": ["Asia"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises from Hong Kong to Asia",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "I would like to see the cruises with a duration between 7 and 14 days",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": 7,"maxDuration": 14,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises between 7 and 14 days",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Show me some cruises that start and end at the same port",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are round-trip cruises",'
            '"round_trip": true,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find cruises where the itinerary starts on or after May 27, 2025",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2025-05-27","maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises starting from May 27, 2025",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Are there any cruises that travel to any European destination",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Europe"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises to Europe",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Can you find for me cruises that include Athens",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Athens"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises that visit Athens",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find all cruises departing from New York that cost less than $18000",
        "Expected Features Extraction": (
            '[{"embarkationPort": ["New York"],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": 18000,"minPrice": null,'
            '"message": "Here are cruises from New York under $18000",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "List all cruises going to South America for under 10000$",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["South America"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": 10000,"minPrice": null,'
            '"message": "Here are cruises to South America under $10000",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "I need a cruise to Barcelona, sailing from June 5th to June 20th",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": ["Barcelona"],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-06-05","maxSailStartDate": "2024-06-05",'
            '"minSailEndDate": "2024-06-20","maxSailEndDate": "2024-06-20",'
            '"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises to Barcelona from June 5th to June 20th",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find a cruise departing from Singapore that finishes before October 26th and stop at Benoa",
        "Expected Features Extraction": (
            '[{"embarkationPort": ["Singapore"],"disembarkationPort": [],'
            '"destinations": ["Benoa"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": "2024-10-26","maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises from Singapore to Benoa ending before October 26th",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find cruises to Kimberley departing from Darwin between June 4th and July 17th",
        "Expected Features Extraction": (
            '[{"embarkationPort": ["Darwin"],"disembarkationPort": [],'
            '"destinations": ["Kimberley"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-06-04","maxSailStartDate": "2024-07-17",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises from Darwin to Kimberley between June 4th and July 17th",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    }, 
    {
        "User Message": "Find cruises that cost less than $5000",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": 5000,"minPrice": null,'
            '"message": "Here are cruises under $5000",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Show me cruises that cost between $2000 and $4000",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": 4000,"minPrice": 2000,'
            '"message": "Here are cruises between $2000 and $4000",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    }, 
    {
        "User Message": "Give me cruises between March and May 2025",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2025-03-01","maxSailStartDate": "2025-05-31",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises between March and May 2025",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Give me cruises are less than 20 days",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": 20,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises less than 20 days",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Give me cruises departing after 13 may 2025",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2025-05-13","maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises departing after May 13, 2025",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "I would like to see the cruises that end on 21 March 2025",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": "2025-03-21","maxSailEndDate": "2025-03-21",'
            '"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises ending on March 21, 2025",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Give me cruises that don't pass by Manaus",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises that avoid Manaus",'
            '"round_trip": false,"ignore_destinations": ["Manaus"],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Show me cruises ending in Vancouver that start in April and cost less than $22200",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": ["Vancouver"],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-04-01","maxSailStartDate": "2024-04-30",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": 22200,"minPrice": null,'
            '"message": "Here are cruises to Vancouver starting in April under $22,200",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Find a cruise with a Mediterranean itinerary, sailing between November 1st and December 15th, costing under $19,000",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Mediterranean"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-11-01","maxSailStartDate": "2024-12-15",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": 19000,"minPrice": null,'
            '"message": "Here are Mediterranean cruises between November 1st and December 15th under $19,000",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Show me cruises to the Transoceanic that sail in May and include stops in Mindelo or Qaqortoq",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Transoceanic", "Mindelo", "Qaqortoq"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-05-01","maxSailStartDate": "2024-05-31",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are Transoceanic cruises in May visiting Mindelo or Qaqortoq",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "List cruises that end in Tokyo, start in March",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": ["Tokyo"],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": "2024-03-01","maxSailStartDate": "2024-03-31",'
            '"minSailEndDate": null,"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are cruises to Tokyo starting in March",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "Show me all discounted cruises available",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": [],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are all cruises with special discounts",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": true,'
            '"is_expedition": false}]'
        )
    },
    {
        "User Message": "I'm looking for expedition cruises to Antarctica",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Antarctica"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are expedition cruises to Antarctica",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": true}]'
        )
    },
    {
        "User Message": "Show me expedition voyages to the Arctic",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Arctic"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are expedition cruises to the Arctic",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": true}]'
        )
    },
    {
        "User Message": "Find expedition-style cruises to the Galapagos Islands",
        "Expected Features Extraction": (
            '[{"embarkationPort": [],"disembarkationPort": [],'
            '"destinations": ["Galapagos Islands"],"minDuration": null,"maxDuration": null,'
            '"minSailStartDate": null,"maxSailStartDate": null,"minSailEndDate": null,'
            '"maxSailEndDate": null,"maxPrice": null,"minPrice": null,'
            '"message": "Here are expedition cruises to the Galapagos Islands",'
            '"round_trip": false,"ignore_destinations": [],"price_discount": false,'
            '"is_expedition": true}]'
        )
    },
]


# Create the sheet if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    df = pd.DataFrame(columns=columns)
    df.to_excel(EXCEL_FILE, index=False)
    
    df = pd.DataFrame(sample_data)
    df.to_excel(EXCEL_FILE, index=False, engine="openpyxl")
    print(f"Created Excel file: {EXCEL_FILE}")

base_url = "http://localhost:5001/api/cruise_search_test"


def get_response(input_text):
    """Send a message to the API and get the response."""
    payload = {
        "sessionId": "eb94989e-f1ee-44f6-b916-bf810281d3f6",
        "message": input_text,
    }
    try:
        response = requests.post(base_url, json=payload)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error making request: {e}")
        return {"error": str(e)}



def validate_features(expected_extraction, actual_extraction):
    """
    Compare expected and actual feature extractions and calculate accuracy scores.
    Returns a dictionary of scores (0 or 1) for each attribute.
    """
    try:
        # Parse expected extraction (which is a string containing a list)
        expected = json.loads(expected_extraction)[0]
        
        # Handle actual extraction (which is already a dict)
        actual = actual_extraction if isinstance(actual_extraction, dict) else json.loads(actual_extraction)[0]
        
        scores = {}
        
        for attr in json_attributes:
            try:
                # Check if both have the attribute
                if attr not in expected or attr not in actual:
                    scores[attr] = 0
                    continue
                
                exp_val = expected[attr]
                act_val = actual[attr]
                
                # Convert None to appropriate empty value based on expected type
                if act_val is None:
                    if isinstance(exp_val, list):
                        act_val = []
                    elif isinstance(exp_val, (int, float)):
                        act_val = None
                    elif isinstance(exp_val, bool):
                        act_val = False
                    elif isinstance(exp_val, str):
                        act_val = ""
                
                # Handle different types of attributes
                if isinstance(exp_val, list):
                    # Ensure we have lists to compare
                    exp_set = set(str(x).lower() for x in (exp_val or []))
                    act_set = set(str(x).lower() for x in (act_val or []))
                    scores[attr] = 1 if exp_set == act_set else 0
                    
                elif isinstance(exp_val, (int, float)) or exp_val is None:
                    # For numeric/null values, check exact match
                    scores[attr] = 1 if exp_val == act_val else 0
                    
                elif isinstance(exp_val, bool):
                    # For boolean values
                    scores[attr] = 1 if exp_val == act_val else 0
                    
                elif isinstance(exp_val, str):
                    # For string values, compare case-insensitive
                    scores[attr] = 1 if str(exp_val or "").lower() == str(act_val or "").lower() else 0
                    
                else:
                    scores[attr] = 0
                
            except Exception as e:
                print(f"Error processing attribute {attr}: {e}")
                scores[attr] = 0
                
        return scores
        
    except Exception as e:
        print(f"Error in validation: {str(e)}")
        return {attr: 0 for attr in json_attributes}

def update_excel():
    """Read input from Excel, send API request, update results, and format sheet for better readability."""
    # Create a new DataFrame with the sample data
    df = pd.DataFrame(sample_data)
    
    # Add all required columns if they don't exist
    for col in columns:
        if col not in df.columns:
            df[col] = None

    for index, row in df.iterrows():
        input_text = row["User Message"]
        expected_extraction = row["Expected Features Extraction"]

        # Get API response
        api_response = get_response(input_text)
        print(f"Row {index} Response done")

        # Calculate scores
        scores = validate_features(expected_extraction, api_response)
        
        # Store the API response as JSON string
        df.loc[index, "Actual Features Extraction"] = json.dumps([api_response])
        
        # Update individual attribute scores
        for attr in json_attributes:
            df.loc[index, f"Score for {attr}"] = scores.get(attr, 0)
        
        # Calculate and update final score
        df.loc[index, "Final Score"] = sum(scores.values()) / len(scores) if scores else 0

    try:
        # Calculate means for score columns and final score
        means = {}
        for col in df.columns:
            if col.startswith("Score for") or col == "Final Score":
                means[col] = df[col].mean()

        # Add summary row
        summary_row = pd.DataFrame([{
            "User Message": "AVERAGE SCORES",
            "Expected Features Extraction": "",
            "Actual Features Extraction": "",
            **means
        }])
        
        # Concatenate the summary row to the DataFrame
        df = pd.concat([df, summary_row], ignore_index=True)

        # Save updates to Excel
        df.to_excel(EXCEL_FILE, index=False, engine="openpyxl")

        # Format Excel file
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Cap the width at 50 to prevent too wide columns
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Enable text wrapping for all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        # Format the summary row in bold
        last_row = ws.max_row
        from openpyxl.styles import Font
        bold_font = Font(bold=True)
        for cell in ws[last_row]:
            cell.font = bold_font

        # Save the formatted file
        wb.save(EXCEL_FILE)
        print("Excel file updated and formatted successfully!")

    except PermissionError:
        print("Error: Please close the Excel file before running this script.")


if __name__ == "__main__":
    update_excel()
